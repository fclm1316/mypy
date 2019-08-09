#!/usr/bin/python3
#coding:utf-8
import openpyxl


file1 = 'C:/Users/DC_JL/Desktop/青海/123/网点信息表.xlsx'
file2 = 'C:/Users/DC_JL/Desktop/青海/123/地区信息.xlsx'
outfile = 'C:/Users/DC_JL/Desktop/青海/123/网点信息表1.xlsx'


workbook1 = openpyxl.load_workbook(file1)
workbook2 = openpyxl.load_workbook(file2)

sheet_name1 = workbook1.sheetnames[1]
sheet_name2 = workbook2.sheetnames[0]

sheet1 = workbook1[sheet_name1]
sheet2 = workbook2[sheet_name2]


nrows1 = sheet1.max_row
nrows2 = sheet2.max_row
dict_1 ={}
for i in range(2,nrows2-1):
    value2_1 = sheet2.cell(row=i,column=1).value
    value2_2 = sheet2.cell(row=i,column=2).value
    dict_1[value2_1]=value2_2

# print(dict_1)

for i in range(2,nrows1 +1):
    value1 = sheet1.cell(row=i,column=7).value
    value2 = sheet1.cell(row=i,column=6)
    for key,values in dict_1.items():
        if value1 == values:
            # print(key,values)
            value2.value=key
            # print(value2)

workbook1.save(outfile)

if __name__ == "__main__":

    pass