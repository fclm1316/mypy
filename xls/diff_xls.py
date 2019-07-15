#!/usr/bin/python3
#coding:utf-8
import openpyxl , xlrd
from openpyxl.styles import Font,colors,fills
from openpyxl.comments import Comment
import os,sys

def diff_xlsx(input_file1,input_file2,outout_file):

    # input_file1 = 'C:/Users/DC_JL/Desktop/c1.xlsx'
    # input_file2 = 'C:/Users/DC_JL/Desktop/c2.xlsx'

    # outout_file = 'C:/Users/DC_JL/Desktop/c3.xlsx'

    workbook1 = openpyxl.load_workbook(input_file1)
    workbook2 = openpyxl.load_workbook(input_file2)

    sheet_name1 = workbook1.sheetnames[0]
    sheet_name2 = workbook2.sheetnames[0]

    sheet1 = workbook1[sheet_name1]
    sheet2 = workbook2[sheet_name2]

    nrows1 = sheet1.max_row
    ncols1 = sheet1.max_column

    nrows2 = sheet2.max_row
    ncols2 = sheet2.max_column
    if nrows1 != nrows2 or ncols1 !=ncols2:
        print(nrows1,nrows2)
        print(ncols1,ncols2)
        print('------')
        print('row or cols not same')
        sys.exit()

    #字体色彩
    # ft = Font(color=colors.YELLOW)
    #背景色彩
    bg_color = fills.GradientFill(stop=['FFFF00','FFFF00'])

    for i in range(1,nrows1+1):
        for j in range(1,ncols1+1):
            # print(i,j)
            value1 = sheet1.cell(row=i,column=j).value
            value2 = sheet2.cell(row=i,column=j).value
            my_comment = Comment('{}'.format(value2),u'fb')
            if value1 != value2:
                # print(sheet1.cell(row=i,column=j).value,sheet2.cell(row=i,column=j).value,i,j)
                # print('----------')
                sheet1.cell(row=i,column=j).comment = my_comment
                # sheet1.cell(row=i,column=j).font = ft
                sheet1.cell(row=i,column=j).fill = bg_color

    workbook1.save(outout_file)



if __name__ == "__main__":
    input_path = 'C:/Users/DC_JL/Desktop/201906/'
    save_path = 'C:/Users/DC_JL/Desktop/201906/new/'
    # file1 = sys.argv[1]
    # file2 = sys.argv[2]
    file1 = os.path.join(input_path,'2019年1-6月柜面业务分离专项考核指标完成情况.xlsx')
    file2 = os.path.join(input_path,'2019年1-6月柜面业务分离专项考核指标完成情况out.xlsx')
    file3 = os.path.join(save_path,'2019年1-6月柜面业务分离专项考核指标完成情况5.xlsx')
    diff_xlsx(file1,file2,file3)
