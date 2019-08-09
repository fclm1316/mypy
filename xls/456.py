#!/usr/bin/python3
#coding:utf-8
import openpyxl
from openpyxl.styles import Font,colors,fills

outfile = 'C:/Users/DC_JL/Desktop/1.xlsx'
outfile2 = 'C:/Users/DC_JL/Desktop/2.xlsx'
file = 'C:/Users/DC_JL/Desktop/base_branchinfo.xlsx'
file2 = 'C:/Users/DC_JL/Desktop/1.xlsx'

workbook = openpyxl.load_workbook(file)
sheet1 = workbook['name1']
sheet2 = workbook['name2']
nrows1 = sheet1.max_row
nrows2 = sheet2.max_row

bg_color = fills.GradientFill(stop=['FFFF00','FFFF00'])

print(nrows1,nrows2)
t = 0
p = 0
for i in range(2,nrows1):
    values1 = sheet1.cell(row = i, column = 2).value
    for j in range(2,nrows2):
        values2 = sheet2.cell(row = j ,column = 2).value
        if str(values1) == str(values2):
            sheet1.cell(row=i,column=2).fill = bg_color
            t = t + 1

# workbook.save(outfile)
print(t)
workbook.save(outfile)


