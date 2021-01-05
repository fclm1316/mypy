#encoding:utf-8
# File: 111.py
# Author:fbi
# Time: 20/12/15
import openpyxl
file = 'C:/Users/DC_JL/Desktop/晟泽泰-外包考勤统计模板2020年11月.xlsx'
workbook = openpyxl.load_workbook(file)
sheet_name = workbook.sheetnames[0]
sheet = workbook[sheet_name]
rows = sheet.max_row
cols = sheet.max_column

a = [16,17,23,24,30,31]

for i in range(3,rows+1):
    add_sum = 0
    add_aa = 0
    for j in range(3,33):
        if j not in a :
            value1 = sheet.cell(row=i, column=j).value
            if value1 is not None:
                jiaban = float(value1) - 8
                add_sum =  add_sum + jiaban
        else:
            value1 = sheet.cell(row=i, column=j).value
            if value1 is not None:
                jiaban2 = float(value1) - 8
                add_aa = add_aa + jiaban2
        sheet.cell(i,34,add_sum)
        sheet.cell(i,35,add_aa)


workbook.save('C:/Users/DC_JL/Desktop/1.xlxs')
